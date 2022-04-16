import numpy as np
import pandas as pd
import openpyxl
from openpyxl.styles import Alignment
from collections import Counter

df=pd.read_excel('table.xlsx','Sheet1',header=None)
a = df.to_numpy()
df=pd.DataFrame(a)
newlst = []
df.applymap(lambda x: newlst.append(x))
newlst.sort()

cou=Counter(newlst)
marks=list(cou.keys())
freq=list(cou.values())

total = len(newlst)

df1 = pd.DataFrame()
df1['marks']=marks
df1['Number of students (i.e.,the frequency)']=freq
df1.to_excel('frequency.xlsx',index=False)

wb=openpyxl.load_workbook("frequency.xlsx")
ws=wb['Sheet1']
ws['A15']='Total'
ws['B15']=total
ws.column_dimensions['B'].width = 25
ws.row_dimensions[1].height = 30
ws['B1'].alignment = Alignment(wrap_text=True)

wb.save("frequency.xlsx")
wb.close()